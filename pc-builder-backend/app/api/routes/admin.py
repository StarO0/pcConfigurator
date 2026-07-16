from __future__ import annotations

import asyncio
import csv
import io
import json
import os
import shutil
import tempfile
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from uuid import UUID
from xml.etree import ElementTree

from fastapi import APIRouter, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import FileResponse
from sqlalchemy import func, or_, select, update
from sqlalchemy.engine import make_url
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload
from sqlalchemy.orm.exc import StaleDataError
from starlette.background import BackgroundTask

from app.api.deps import AdminUser, DbSession, OptionalUser, ServiceAuth
from app.core.config import settings
from app.core.security import generate_opaque_token, token_hash
from app.models.entities import (
    AuditLog,
    Build,
    BuildComponent,
    Offer,
    ParserRun,
    Product,
    ServiceToken,
    Store,
    User,
)
from app.schemas.admin import (
    AdminStatsOut,
    AdminUserUpdate,
    AuditLogOut,
    DuplicateGroupOut,
    DuplicateProductOut,
    FilePreviewResponse,
    LocalRestoreResponse,
    MergeProductsRequest,
    NormalizationResponse,
    ParserRunOut,
    ServiceTokenCreate,
    ServiceTokenCreated,
    TaskQueuedResponse,
)
from app.schemas.auth import UserOut
from app.schemas.common import MessageResponse, Page
from app.schemas.products import (
    OfferImportRequest,
    OfferImportResponse,
    ProductCreate,
    ProductImportRequest,
    ProductImportResponse,
    ProductOut,
    ProductUpdate,
    StoreCreate,
    StoreOut,
    StoreUpdate,
)
from app.services.audit import audit
from app.services.matching import normalize_text
from app.services.offer_import import import_offers
from app.services.parsers.sync import sync_store
from app.services.serializers import product_to_schema, store_to_schema
from app.services.spec_normalization import normalize_specs

router = APIRouter(prefix="/admin", tags=["admin"])

MAX_PREVIEW_ROWS = 100_000
MAX_IMPORT_FILE_SIZE = 25_000_000


def _normalized_column(value: object) -> str:
    raw = str(value or "").strip().lower()
    result: list[str] = []
    previous_separator = False
    for character in raw:
        if character.isalnum() or character in "а-яё":
            result.append(character)
            previous_separator = False
        elif not previous_separator:
            result.append("_")
            previous_separator = True
    return "".join(result).strip("_")


def _json_value(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _normalize_rows(rows: list[dict[object, object]]) -> list[dict[str, object]]:
    normalized: list[dict[str, object]] = []
    for row in rows[:MAX_PREVIEW_ROWS]:
        item = {
            column: _json_value(value)
            for raw_column, value in row.items()
            if (column := _normalized_column(raw_column))
        }
        if any(value not in (None, "") for value in item.values()):
            normalized.append(item)
    return normalized


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1251", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Не удалось определить кодировку файла")


def _xml_tag(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _flatten_xml_element(element: ElementTree.Element, prefix: str = "") -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in element.attrib.items():
        result[f"{prefix}{_xml_tag(key)}"] = value

    children = list(element)
    if not children:
        value = (element.text or "").strip()
        if prefix and value:
            result[prefix.rstrip("_")] = value
        return result

    grouped: dict[str, list[ElementTree.Element]] = {}
    for child in children:
        grouped.setdefault(_xml_tag(child.tag), []).append(child)

    for tag, items in grouped.items():
        key_prefix = f"{prefix}{tag}"
        if len(items) == 1:
            child = items[0]
            if list(child) or child.attrib:
                result.update(_flatten_xml_element(child, f"{key_prefix}_"))
            else:
                value = (child.text or "").strip()
                if value:
                    result[key_prefix] = value
        else:
            values: list[str] = []
            for child in items:
                if list(child) or child.attrib:
                    flattened = _flatten_xml_element(child)
                    values.append(json.dumps(flattened, ensure_ascii=False))
                else:
                    value = (child.text or "").strip()
                    if value:
                        values.append(value)
            if values:
                result[key_prefix] = values
    return result


def _parse_xml_rows(data: bytes) -> list[dict[object, object]]:
    upper = data[:4096].upper()
    if b"<!DOCTYPE" in upper or b"<!ENTITY" in upper:
        raise ValueError("XML с DOCTYPE/ENTITY не поддерживается")
    root = ElementTree.fromstring(data)
    children = list(root)
    if not children:
        return [_flatten_xml_element(root)]

    row_names = {"product", "offer", "item", "row", "record"}
    candidates = [child for child in children if _xml_tag(child.tag).lower() in row_names]
    if not candidates and len(children) == 1:
        nested = list(children[0])
        candidates = [child for child in nested if _xml_tag(child.tag).lower() in row_names]
        if not candidates:
            candidates = nested
    if not candidates:
        candidates = children
    return [_flatten_xml_element(element) for element in candidates]


def _parse_import_file(filename: str, data: bytes) -> tuple[list[dict[str, object]], int]:
    suffix = Path(filename).suffix.lower()
    raw_rows: list[dict[object, object]]
    if suffix == ".json":
        payload = json.loads(_decode_text(data))
        if isinstance(payload, dict):
            payload = payload.get("products") or payload.get("offers")
        if not isinstance(payload, list) or not all(isinstance(item, dict) for item in payload):
            raise ValueError("JSON должен содержать массив объектов, products[] или offers[]")
        raw_rows = payload
    elif suffix == ".csv":
        sample = _decode_text(data)
        try:
            dialect = csv.Sniffer().sniff(sample[:8192], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        raw_rows = list(csv.DictReader(io.StringIO(sample), dialect=dialect))
    elif suffix == ".xml":
        raw_rows = _parse_xml_rows(data)
    elif suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            return [], 0
        raw_rows = [dict(zip(headers, values, strict=False)) for values in iterator]
        workbook.close()
    elif suffix == ".xls":
        raise ValueError("Старый .xls не поддерживается: сохраните файл как .xlsx или .csv")
    else:
        raise ValueError("Поддерживаются CSV, XLSX, XLSM, JSON и XML")
    total = len(raw_rows)
    return _normalize_rows(raw_rows), total


def require_import_auth(user, service_token) -> None:
    if user and user.role == "admin":
        return
    if service_token and "offers:write" in service_token.scopes:
        return
    raise HTTPException(
        status_code=403, detail="Требуются права admin или service token offers:write"
    )


@router.get("/stats", response_model=AdminStatsOut)
async def admin_stats(session: DbSession, _: AdminUser) -> AdminStatsOut:
    stale_before = datetime.now(UTC) - timedelta(hours=settings.offer_stale_hours)
    failed_since = datetime.now(UTC) - timedelta(hours=24)
    return AdminStatsOut(
        users=await session.scalar(select(func.count(User.id))) or 0,
        products=await session.scalar(
            select(func.count(Product.id)).where(Product.is_active.is_(True))
        )
        or 0,
        active_offers=await session.scalar(
            select(func.count(Offer.id)).where(Offer.is_active.is_(True), Offer.in_stock.is_(True))
        )
        or 0,
        saved_builds=await session.scalar(
            select(func.count(Build.id)).where(Build.is_saved.is_(True), Build.deleted_at.is_(None))
        )
        or 0,
        stale_offers=await session.scalar(
            select(func.count(Offer.id)).where(
                Offer.fetched_at < stale_before, Offer.is_active.is_(True)
            )
        )
        or 0,
        failed_parser_runs_24h=await session.scalar(
            select(func.count(ParserRun.id)).where(
                ParserRun.status == "failed", ParserRun.started_at >= failed_since
            )
        )
        or 0,
    )


@router.post("/products", response_model=ProductOut, status_code=status.HTTP_201_CREATED)
async def create_product(
    payload: ProductCreate, session: DbSession, request: Request, admin: AdminUser
) -> ProductOut:
    product = Product(
        category=payload.category.lower(),
        brand=payload.brand.strip(),
        name=payload.name.strip(),
        normalized_name=normalize_text(f"{payload.brand} {payload.name} {payload.sku}"),
        sku=payload.sku.strip(),
        ean=payload.ean,
        mpn=payload.mpn,
        image_url=str(payload.image_url) if payload.image_url else None,
        gallery_urls=[str(url) for url in payload.gallery_urls],
        canonical_source=payload.canonical_source,
        canonical_id=payload.canonical_id,
        source_url=str(payload.source_url) if payload.source_url else None,
        release_date=payload.release_date,
        performance_score=payload.performance_score,
        noise_score=payload.noise_score,
        upgrade_score=payload.upgrade_score,
        quality_score=payload.quality_score,
        specs=normalize_specs(
            payload.category.lower(), payload.name.strip(), payload.brand.strip(), payload.specs
        ),
        status=payload.status,
        is_active=payload.status == "active",
    )
    session.add(product)
    try:
        await session.flush()
    except IntegrityError as exc:
        raise HTTPException(status_code=409, detail="SKU/EAN уже существует") from exc
    await audit(session, request, "admin.product_create", "product", product.id, user_id=admin.id)
    await session.commit()
    result = await session.execute(
        select(Product)
        .options(
            selectinload(Product.offers).selectinload(Offer.store),
            selectinload(Product.benchmarks),
        )
        .where(Product.id == product.id)
    )
    return product_to_schema(result.scalar_one())


@router.post("/product-tools/import", response_model=ProductImportResponse)
async def import_products(
    payload: ProductImportRequest,
    session: DbSession,
    request: Request,
    admin: AdminUser,
) -> ProductImportResponse:
    created = 0
    updated_count = 0
    skipped = 0
    errors: list[str] = []
    skus = {item.sku for item in payload.products}
    eans = {item.ean for item in payload.products if item.ean}
    mpns = {item.mpn for item in payload.products if item.mpn}
    filters = [Product.sku.in_(skus)]
    if eans:
        filters.append(Product.ean.in_(eans))
    if mpns:
        filters.append(Product.mpn.in_(mpns))
    existing_products = (await session.execute(select(Product).where(or_(*filters)))).scalars()
    by_sku: dict[str, Product] = {}
    by_ean: dict[str, Product] = {}
    by_mpn_brand: dict[tuple[str, str], Product] = {}
    for product in existing_products:
        by_sku[product.sku] = product
        if product.ean:
            by_ean[product.ean] = product
        if product.mpn:
            by_mpn_brand[(product.mpn, product.brand.casefold())] = product

    for item in payload.products:
        existing = by_sku.get(item.sku)
        if existing is None and item.ean:
            existing = by_ean.get(item.ean)
        if existing is None and item.mpn:
            existing = by_mpn_brand.get((item.mpn, item.brand.casefold()))
        normalized = normalize_specs(item.category.lower(), item.name, item.brand, item.specs)
        if existing:
            if not payload.update_existing:
                skipped += 1
                continue
            existing.category = item.category.lower()
            existing.brand = item.brand.strip()
            existing.name = item.name.strip()
            existing.normalized_name = normalize_text(f"{item.brand} {item.name} {item.mpn or ''}")
            existing.ean = item.ean or existing.ean
            existing.mpn = item.mpn or existing.mpn
            existing.image_url = str(item.image_url) if item.image_url else existing.image_url
            if item.gallery_urls:
                existing.gallery_urls = [str(url) for url in item.gallery_urls]
            existing.canonical_source = item.canonical_source or existing.canonical_source
            existing.canonical_id = item.canonical_id or existing.canonical_id
            existing.source_url = str(item.source_url) if item.source_url else existing.source_url
            if item.performance_score is not None:
                existing.performance_score = item.performance_score
            existing.specs = {**(existing.specs or {}), **normalized}
            existing.status = "active"
            existing.is_active = True
            updated_count += 1
        else:
            product = Product(
                category=item.category.lower(),
                brand=item.brand.strip(),
                name=item.name.strip(),
                normalized_name=normalize_text(f"{item.brand} {item.name} {item.mpn or ''}"),
                sku=item.sku,
                ean=item.ean,
                mpn=item.mpn,
                image_url=str(item.image_url) if item.image_url else None,
                gallery_urls=[str(url) for url in item.gallery_urls],
                canonical_source=item.canonical_source,
                canonical_id=item.canonical_id,
                source_url=str(item.source_url) if item.source_url else None,
                performance_score=item.performance_score or 0,
                specs=normalized,
                status="active",
                is_active=True,
            )
            session.add(product)
            by_sku[item.sku] = product
            if item.ean:
                by_ean[item.ean] = product
            if item.mpn:
                by_mpn_brand[(item.mpn, item.brand.casefold())] = product
            created += 1
    try:
        await audit(
            session,
            request,
            "admin.products_import",
            "product",
            user_id=admin.id,
            details={"created": created, "updated": updated_count, "skipped": skipped},
        )
        await session.commit()
    except IntegrityError as exc:
        await session.rollback()
        raise HTTPException(
            status_code=409, detail=f"Конфликт SKU/EAN в импортируемом файле: {exc.orig}"
        ) from exc
    return ProductImportResponse(
        created=created,
        updated=updated_count,
        skipped=skipped,
        errors=errors[:200],
    )


@router.post("/file-preview", response_model=FilePreviewResponse)
async def preview_import_file(
    _: AdminUser,
    file: UploadFile = File(...),
) -> FilePreviewResponse:
    filename = file.filename or "import"
    data = await file.read(MAX_IMPORT_FILE_SIZE + 1)
    if len(data) > MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=413, detail="Файл импорта больше 25 MB")
    try:
        rows, total = await asyncio.to_thread(_parse_import_file, filename, data)
    except (csv.Error, json.JSONDecodeError, ElementTree.ParseError, OSError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return FilePreviewResponse(
        rows=rows,
        total=total,
        truncated=total > MAX_PREVIEW_ROWS,
    )


@router.patch("/products/{product_id}", response_model=ProductOut)
async def update_product(
    product_id: UUID, payload: ProductUpdate, session: DbSession, request: Request, admin: AdminUser
) -> ProductOut:
    result = await session.execute(
        select(Product)
        .options(
            selectinload(Product.offers).selectinload(Offer.store), selectinload(Product.benchmarks)
        )
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if product is None:
        raise HTTPException(status_code=404, detail="Товар не найден")
    if product.version != payload.version:
        raise HTTPException(
            status_code=409,
            detail={"message": "Товар уже изменён", "current_version": product.version},
        )
    values = payload.model_dump(exclude_unset=True, exclude={"version"})
    if "image_url" in values and values["image_url"] is not None:
        values["image_url"] = str(values["image_url"])
    if "gallery_urls" in values and values["gallery_urls"] is not None:
        values["gallery_urls"] = [str(url) for url in values["gallery_urls"]]
    if "source_url" in values and values["source_url"] is not None:
        values["source_url"] = str(values["source_url"])
    for key, value in values.items():
        setattr(product, key, value)
    if "name" in values or "brand" in values:
        product.normalized_name = normalize_text(f"{product.brand} {product.name} {product.sku}")
    if {"name", "brand", "specs"} & values.keys():
        product.specs = normalize_specs(
            product.category, product.name, product.brand, product.specs or {}
        )
    if product.status != "active":
        product.is_active = False
    await audit(
        session,
        request,
        "admin.product_update",
        "product",
        product.id,
        user_id=admin.id,
        details={"fields": list(values)},
    )
    try:
        await session.commit()
    except StaleDataError as exc:
        raise HTTPException(status_code=409, detail="Товар изменён параллельно") from exc
    return product_to_schema(product)


@router.delete("/products/{product_id}", response_model=MessageResponse)
async def deactivate_product(
    product_id: UUID, session: DbSession, request: Request, admin: AdminUser
) -> MessageResponse:
    product = await session.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="Товар не найден")
    product.is_active = False
    product.status = "discontinued"
    await audit(
        session, request, "admin.product_deactivate", "product", product.id, user_id=admin.id
    )
    await session.commit()
    return MessageResponse(message="Товар отключён")


@router.post("/product-tools/merge", response_model=MessageResponse)
async def merge_products(
    payload: MergeProductsRequest, session: DbSession, request: Request, admin: AdminUser
) -> MessageResponse:
    if payload.source_product_id == payload.target_product_id:
        raise HTTPException(status_code=400, detail="Нужны разные товары")
    source = await session.get(Product, payload.source_product_id)
    target = await session.get(Product, payload.target_product_id)
    if source is None or target is None:
        raise HTTPException(status_code=404, detail="Один из товаров не найден")
    await session.execute(
        update(Offer).where(Offer.product_id == source.id).values(product_id=target.id)
    )
    await session.execute(
        update(BuildComponent)
        .where(BuildComponent.product_id == source.id)
        .values(product_id=target.id)
    )
    source.is_active = False
    source.status = "merged"
    await audit(
        session,
        request,
        "admin.product_merge",
        "product",
        target.id,
        user_id=admin.id,
        details={"source": str(source.id)},
    )
    await session.commit()
    return MessageResponse(message="Товары объединены")


@router.get("/product-tools/duplicates", response_model=list[DuplicateGroupOut])
async def duplicate_products(
    session: DbSession,
    _: AdminUser,
    limit: int = Query(default=50, ge=1, le=200),
) -> list[DuplicateGroupOut]:
    keys = (
        await session.execute(
            select(Product.normalized_name, func.count(Product.id))
            .where(Product.is_active.is_(True), Product.normalized_name != "")
            .group_by(Product.normalized_name)
            .having(func.count(Product.id) > 1)
            .order_by(func.count(Product.id).desc(), Product.normalized_name)
            .limit(limit)
        )
    ).all()
    groups: list[DuplicateGroupOut] = []
    for key, _count in keys:
        products = (
            await session.execute(
                select(Product)
                .where(Product.is_active.is_(True), Product.normalized_name == key)
                .order_by(Product.created_at)
                .limit(20)
            )
        ).scalars()
        groups.append(
            DuplicateGroupOut(
                key=key,
                reason="normalized_name",
                products=[
                    DuplicateProductOut(
                        id=item.id,
                        category=item.category,
                        brand=item.brand,
                        name=item.name,
                        sku=item.sku,
                        ean=item.ean,
                        mpn=item.mpn,
                    )
                    for item in products
                ],
            )
        )
    return groups


@router.post("/product-tools/normalize", response_model=NormalizationResponse)
async def normalize_products(
    session: DbSession,
    request: Request,
    admin: AdminUser,
    dry_run: bool = False,
    limit: int = Query(default=100_000, ge=1, le=200_000),
) -> NormalizationResponse:
    scanned = 0
    changed = 0
    batch_size = 250
    for offset in range(0, limit, batch_size):
        result = await session.execute(
            select(Product).order_by(Product.id).offset(offset).limit(batch_size)
        )
        products = list(result.scalars())
        if not products:
            break
        for product in products:
            scanned += 1
            normalized = normalize_specs(
                product.category, product.name, product.brand, product.specs or {}
            )
            if normalized != product.specs:
                changed += 1
                if not dry_run:
                    product.specs = normalized
        if not dry_run:
            await session.commit()
    await audit(
        session,
        request,
        "admin.products_normalize",
        "product",
        user_id=admin.id,
        details={"scanned": scanned, "changed": changed, "dry_run": dry_run},
    )
    await session.commit()
    return NormalizationResponse(scanned=scanned, changed=changed, dry_run=dry_run)


@router.get("/stores", response_model=list[StoreOut])
async def list_stores(session: DbSession, _: AdminUser) -> list[StoreOut]:
    result = await session.execute(select(Store).order_by(Store.name))
    return [store_to_schema(item) for item in result.scalars()]


@router.post("/stores", response_model=StoreOut, status_code=status.HTTP_201_CREATED)
async def create_store(
    payload: StoreCreate, session: DbSession, request: Request, admin: AdminUser
) -> StoreOut:
    store = Store(
        slug=payload.slug,
        name=payload.name,
        base_url=str(payload.base_url),
        country=payload.country.upper(),
        parser_type=payload.parser_type,
        parser_config=payload.parser_config,
    )
    session.add(store)
    try:
        await session.flush()
    except IntegrityError as exc:
        raise HTTPException(
            status_code=409, detail="Магазин с таким slug/name уже существует"
        ) from exc
    await audit(session, request, "admin.store_create", "store", store.id, user_id=admin.id)
    await session.commit()
    return store_to_schema(store)


@router.patch("/stores/{store_id}", response_model=StoreOut)
async def update_store(
    store_id: UUID, payload: StoreUpdate, session: DbSession, request: Request, admin: AdminUser
) -> StoreOut:
    store = await session.get(Store, store_id)
    if store is None:
        raise HTTPException(status_code=404, detail="Магазин не найден")
    values = payload.model_dump(exclude_unset=True)
    if "base_url" in values and values["base_url"] is not None:
        values["base_url"] = str(values["base_url"])
    for key, value in values.items():
        setattr(store, key, value)
    await audit(session, request, "admin.store_update", "store", store.id, user_id=admin.id)
    await session.commit()
    return store_to_schema(store)


@router.post("/offers/import", response_model=OfferImportResponse)
async def import_offer_feed(
    payload: OfferImportRequest,
    session: DbSession,
    request: Request,
    user: OptionalUser,
    service_token: ServiceAuth,
) -> OfferImportResponse:
    require_import_auth(user, service_token)
    imported = await import_offers(
        session,
        payload.offers,
        create_unmatched_products=payload.create_unmatched_products,
    )
    await audit(
        session,
        request,
        "offers.import",
        "offer",
        user_id=user.id if user else None,
        service_token_id=service_token.id if service_token else None,
        details={
            "created": imported.created,
            "updated": imported.updated,
            "skipped": imported.skipped,
        },
    )
    await session.commit()
    return OfferImportResponse(
        created=imported.created,
        updated=imported.updated,
        skipped=imported.skipped,
        unmatched=imported.unmatched,
    )


@router.post(
    "/service-tokens", response_model=ServiceTokenCreated, status_code=status.HTTP_201_CREATED
)
async def create_service_token(
    payload: ServiceTokenCreate, session: DbSession, request: Request, admin: AdminUser
) -> ServiceTokenCreated:
    raw = generate_opaque_token()
    item = ServiceToken(
        name=payload.name,
        token_hash=token_hash(raw),
        scopes=payload.scopes,
        expires_at=payload.expires_at,
    )
    session.add(item)
    try:
        await session.flush()
    except IntegrityError as exc:
        raise HTTPException(
            status_code=409, detail="Service token с таким именем уже существует"
        ) from exc
    await audit(
        session,
        request,
        "admin.service_token_create",
        "service_token",
        item.id,
        user_id=admin.id,
        details={"scopes": item.scopes},
    )
    await session.commit()
    return ServiceTokenCreated(
        id=item.id, name=item.name, token=raw, scopes=item.scopes, expires_at=item.expires_at
    )


@router.delete("/service-tokens/{token_id}", response_model=MessageResponse)
async def revoke_service_token(
    token_id: UUID, session: DbSession, request: Request, admin: AdminUser
) -> MessageResponse:
    item = await session.get(ServiceToken, token_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Service token не найден")
    item.is_active = False
    await audit(
        session, request, "admin.service_token_revoke", "service_token", item.id, user_id=admin.id
    )
    await session.commit()
    return MessageResponse(message="Service token отозван")


@router.post(
    "/stores/{store_id}/sync",
    response_model=TaskQueuedResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def queue_store_sync(store_id: UUID, session: DbSession, _: AdminUser) -> TaskQueuedResponse:
    store = await session.get(Store, store_id)
    if store is None or not store.is_active:
        raise HTTPException(status_code=404, detail="Активный магазин не найден")
    if store.parser_type == "manual":
        raise HTTPException(status_code=409, detail="У магазина не настроен автоматический parser")
    if store.parser_type != "ceneo" and not (store.parser_config or {}).get("terms_confirmed"):
        raise HTTPException(
            status_code=409,
            detail="Сначала подтвердите robots.txt и условия автоматического обхода источника",
        )
    from app.workers.tasks import sync_store_task

    task = sync_store_task.delay(str(store_id))
    return TaskQueuedResponse(task_id=task.id)


@router.post("/stores/{store_id}/sync-now", response_model=ParserRunOut)
async def sync_store_now(store_id: UUID, session: DbSession, _: AdminUser) -> ParserRunOut:
    store = await session.get(Store, store_id)
    if store is None or not store.is_active:
        raise HTTPException(status_code=404, detail="Активный магазин не найден")
    if store.parser_type == "manual":
        raise HTTPException(status_code=409, detail="У магазина не настроен автоматический parser")
    if store.parser_type != "ceneo" and not (store.parser_config or {}).get("terms_confirmed"):
        raise HTTPException(
            status_code=409,
            detail="Сначала подтвердите robots.txt и условия автоматического обхода источника",
        )
    run = await sync_store(session, store, task_id="local-sync")
    return ParserRunOut.model_validate(run, from_attributes=True)


def _postgres_command(command: str) -> tuple[list[str], dict[str, str]]:
    url = make_url(settings.database_url)
    if url.get_backend_name() != "postgresql":
        raise HTTPException(status_code=409, detail="Backup поддерживается только для PostgreSQL")
    binary = shutil.which(command)
    if binary is None:
        raise HTTPException(
            status_code=503,
            detail=f"{command} не установлен в API-контейнере",
        )
    args = [binary]
    if url.host:
        args.extend(["--host", url.host])
    if url.port:
        args.extend(["--port", str(url.port)])
    if url.username:
        args.extend(["--username", url.username])
    if url.database:
        args.extend(["--dbname", url.database])
    env = os.environ.copy()
    if url.password:
        env["PGPASSWORD"] = url.password
    return args, env


async def _run_postgres_command(args: list[str], env: dict[str, str]) -> None:
    process = await asyncio.create_subprocess_exec(
        *args,
        env=env,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    _stdout, stderr = await process.communicate()
    if process.returncode != 0:
        detail = stderr.decode("utf-8", errors="replace").strip()[-2000:]
        raise HTTPException(status_code=500, detail=f"PostgreSQL utility failed: {detail}")


@router.get("/local-backup", response_class=FileResponse)
async def local_backup(_: AdminUser) -> FileResponse:
    with tempfile.NamedTemporaryFile(
        prefix="pc-builder-postgres-", suffix=".dump", delete=False
    ) as handle:
        target = Path(handle.name)
    args, env = _postgres_command("pg_dump")
    args.extend(["--format=custom", "--no-owner", "--no-acl", "--file", str(target)])
    try:
        await _run_postgres_command(args, env)
    except Exception:
        await asyncio.to_thread(target.unlink, missing_ok=True)
        raise
    filename = f"pc-builder-postgres-{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}.dump"
    return FileResponse(
        target,
        filename=filename,
        media_type="application/octet-stream",
        background=BackgroundTask(target.unlink, missing_ok=True),
    )


@router.post("/local-restore", response_model=LocalRestoreResponse)
async def local_restore(
    _: AdminUser,
    database: UploadFile = File(...),
) -> LocalRestoreResponse:
    data = await database.read(250_000_001)
    if len(data) > 250_000_000:
        raise HTTPException(status_code=413, detail="Backup больше 250 MB")
    with tempfile.NamedTemporaryFile(
        prefix="pc-builder-restore-", suffix=".dump", delete=False
    ) as handle:
        pending = Path(handle.name)
        handle.write(data)
    try:
        list_args, env = _postgres_command("pg_restore")
        await _run_postgres_command([list_args[0], "--list", str(pending)], env)
        restore_args, env = _postgres_command("pg_restore")
        restore_args.extend(
            [
                "--clean",
                "--if-exists",
                "--no-owner",
                "--no-acl",
                "--exit-on-error",
                str(pending),
            ]
        )
        await _run_postgres_command(restore_args, env)
    except HTTPException as exc:
        await asyncio.to_thread(pending.unlink, missing_ok=True)
        raise HTTPException(
            status_code=422,
            detail=f"Не удалось проверить или восстановить PostgreSQL backup: {exc.detail}",
        ) from exc
    await asyncio.to_thread(pending.unlink, missing_ok=True)
    return LocalRestoreResponse(
        message="PostgreSQL backup восстановлен. Перезапустите API и worker."
    )


@router.get("/parser-runs", response_model=Page[ParserRunOut])
async def parser_runs(
    session: DbSession,
    _: AdminUser,
    status_filter: str | None = Query(default=None, alias="status"),
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
) -> Page[ParserRunOut]:
    filters = []
    if status_filter:
        filters.append(ParserRun.status == status_filter)
    total = await session.scalar(select(func.count(ParserRun.id)).where(*filters)) or 0
    result = await session.execute(
        select(ParserRun)
        .where(*filters)
        .order_by(ParserRun.started_at.desc())
        .offset(offset)
        .limit(limit)
    )
    return Page[ParserRunOut](
        items=[
            ParserRunOut.model_validate(item, from_attributes=True) for item in result.scalars()
        ],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.get("/audit-logs", response_model=Page[AuditLogOut])
async def audit_logs(
    session: DbSession,
    _: AdminUser,
    action: str | None = None,
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> Page[AuditLogOut]:
    filters = [AuditLog.action == action] if action else []
    total = await session.scalar(select(func.count(AuditLog.id)).where(*filters)) or 0
    result = await session.execute(
        select(AuditLog)
        .where(*filters)
        .order_by(AuditLog.created_at.desc())
        .offset(offset)
        .limit(limit)
    )
    return Page[AuditLogOut](
        items=[AuditLogOut.model_validate(item, from_attributes=True) for item in result.scalars()],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.get("/users", response_model=Page[UserOut])
async def admin_users(
    session: DbSession,
    _: AdminUser,
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> Page[UserOut]:
    total = await session.scalar(select(func.count(User.id))) or 0
    result = await session.execute(
        select(User).order_by(User.created_at.desc()).offset(offset).limit(limit)
    )
    items = [
        UserOut(
            id=item.id,
            email=item.email,
            display_name=item.display_name,
            role=item.role,
            is_active=item.is_active,
            is_verified=item.is_verified,
            created_at=item.created_at,
        )
        for item in result.scalars()
    ]
    return Page[UserOut](items=items, total=total, limit=limit, offset=offset)


@router.patch("/users/{user_id}", response_model=UserOut)
async def admin_update_user(
    user_id: UUID, payload: AdminUserUpdate, session: DbSession, request: Request, admin: AdminUser
) -> UserOut:
    user = await session.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if user.id == admin.id and payload.is_active is False:
        raise HTTPException(status_code=400, detail="Нельзя отключить собственный аккаунт")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(user, key, value)
    await audit(
        session,
        request,
        "admin.user_update",
        "user",
        user.id,
        user_id=admin.id,
        details=payload.model_dump(exclude_unset=True),
    )
    await session.commit()
    return UserOut(
        id=user.id,
        email=user.email,
        display_name=user.display_name,
        role=user.role,
        is_active=user.is_active,
        is_verified=user.is_verified,
        created_at=user.created_at,
    )
